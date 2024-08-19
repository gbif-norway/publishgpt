from api.models import Dataset, Table, Agent, Message, Task
from rest_framework import serializers
import pandas as pd
import openpyxl
import tempfile
import os
from api.helpers import discord_bot


class TaskSerializer(serializers.ModelSerializer):
    class Meta:
        model = Task
        fields = ['id', 'name', 'per_table', 'attempt_autonomous']


class TableSerializer(serializers.ModelSerializer):
    df_str = serializers.CharField(source='df', read_only=True)

    class Meta:
        model = Table
        fields = ['id', 'created_at', 'updated_at', 'dataset', 'title', 'df_str', 'description', 'df_json']


class TableShortSerializer(serializers.ModelSerializer):
    class Meta:
        model = Table
        fields = ['id', 'title', 'updated_at']


class MessageSerializer(serializers.ModelSerializer):
    role = serializers.SerializerMethodField()

    class Meta:
        model = Message
        fields = '__all__'

    def get_role(self, obj):
        return obj.role  


class AgentSerializer(serializers.ModelSerializer):
    message_set = MessageSerializer(many=True, read_only=True)
    task = TaskSerializer(read_only=True)
    table_set = TableShortSerializer(many=True, read_only=True)

    class Meta:
        model = Agent
        fields = '__all__'


class DatasetSerializer(serializers.ModelSerializer):
    visible_agent_set = serializers.SerializerMethodField()

    class Meta:
        model = Dataset
        fields = '__all__'
    
    def get_visible_agent_set(self, dataset):
        agents = list(dataset.agent_set.filter(completed_at__isnull=False))
        next_active_agent = dataset.agent_set.filter(completed_at__isnull=True).first()
        if next_active_agent:
            agents.append(next_active_agent)
        return AgentSerializer(agents, many=True).data

    def create(self, data):
        discord_bot.send_discord_message(f"New dataset publication starting on ChatIPT. User file: {data['file'].name}.")
        try:
            df = pd.read_csv(data['file'].file, dtype='str', encoding='utf-8', encoding_errors='surrogateescape')
            if len(df) < 4:
                raise serializers.ValidationError(f"Your dataset has only {len(df)} rows, are you sure you uploaded the right thing? I need a larger spreadsheet to be able to help you with publication.")
            dfs = {data['file'].name: df}
        except:
            # dfs = pd.read_excel(data['file'].file, dtype='str', sheet_name=None)
            workbook = openpyxl.load_workbook(data['file'].file)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # 'f' indicates a formula
                            cell.value = '' # f'[FORMULA: {cell.value}]'
                for merged_cell in list(sheet.merged_cells.ranges):
                    min_col, min_row, max_col, max_row = merged_cell.min_col, merged_cell.min_row, merged_cell.max_col, merged_cell.max_row
                    value = sheet.cell(row=min_row, column=min_col).value
                    sheet.unmerge_cells(str(merged_cell))
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            sheet.cell(row=row, column=col).value = f"{value} [UNMERGED CELL]"
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                temp_file_name = tmp.name
                workbook.save(temp_file_name)
            dfs = pd.read_excel(temp_file_name, dtype='str', sheet_name=None)
            os.remove(temp_file_name)

        dataset = Dataset.objects.create(**data)
        tables = []
        for sheet_name, df in dfs.items():
            if not df.empty:
                tables.append(Table.objects.create(dataset=dataset, title=sheet_name, df=df))
        agent = Agent.create_with_system_message(dataset=dataset, task=Task.objects.first(), tables=tables)
        discord_bot.send_discord_message(f"Dataset ID assigned: {dataset.id}.")
        return dataset
